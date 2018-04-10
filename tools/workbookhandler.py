#%%
# -*- encoding: utf8-*-

import os
import openpyxl
import csv
import pandas as pd
import numpy as np
from define import *

class WorkBookHandler:
    @staticmethod
    def load_workbook(src):       
        try:
            if not os.path.isfile(src):
                print("File {} is not exist.".format(src))
                return None
            wb = openpyxl.load_workbook(src)
            wb.suject = src
            return  wb
        except IOError:
            print('File is opened by other process, close it.')
            return None

    @staticmethod
    def remove_sheet(wb, sheet_name):   
        next_active_sheet_index = -1
        if sheet_name in wb:
            for i in range(0, len(wb.sheetnames)):
                if wb.sheetnames[i] != sheet_name:
                    next_active_sheet_index = i
                    print('active sheetname {}'.format(next_active_sheet_index))
                    break
            if next_active_sheet_index == -1:
                print('There just has only one sheet {}, you can\'t remove it'.format(sheet_name))
                return False
            wb.active = i
            del wb[sheet_name]           
            print(wb.sheetnames)
            return True
        else:
            print('{} not contains in the {}'.format(sheet_name, wb.subject))   
            return False     
        
    @staticmethod
    def create_sheet(wb, sheet_name, datas, start_col, start_row, end_col, end_row):
        print('wb: ' + str(wb))
        ws = wb.create_sheet(0)
        ws.title = Define.STOCK_LIST_SHEET_NAME
        WorkBookHandler.paste_datas(ws, datas, start_col, start_row, end_col, end_row)
        print('sheets: {}'.format(wb.sheetnames) )


    @staticmethod
    def copy_datas(ws, startCol, startRow, endCol, endRow):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(ws.cell(row = i, column = j).value)
                #Adds the RowSelected List and nests inside the rangeSelected
                rangeSelected.append(rowSelected) 
        return rangeSelected

    @staticmethod
    def paste_datas(ws, datas, startCol, startRow, endCol, endRow ):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                ws.cell(row = i, column = j).value = datas[countRow][countCol]
                countCol += 1
            countRow += 1

if __name__ == '__main__':
    try:
        print('Begin update')
        #open workbook
        wb = WorkBookHandler.load_workbook(Define.XLS_PATH)

        #remove list id sheet
        WorkBookHandler.remove_sheet(wb, Define.STOCK_LIST_SHEET_NAME)

        #load csv files
        df2 =pd.read_csv(Define.get_list_path(MarketType.TSE), encoding='utf-8', header=None)
        df4 =pd.read_csv(Define.get_list_path(MarketType.OTC), encoding='utf-8', header=None)  

        #delete title key
        nd2=np.delete(df2.values, np.argwhere(df2.values == 'id')[0][0], axis=0) 
        nd4=np.delete(df4.values, np.argwhere(df4.values == 'id')[0][0], axis=0) 

        #combine stock id list 
        arr = np.concatenate((nd2, nd4), axis=0)        
    
        #create new list id sheet
        print('Create sheet')
        WorkBookHandler.create_sheet(wb, Define.STOCK_LIST_SHEET_NAME, arr, 1, 1, len(arr[0]), len(arr))
        
        #In this case, save file twice will throw error, just keep this action at the end.
        wb.save(Define.XLS_PATH)
        wb.close()
    except PermissionError:
         print('File is opened by other process, close it.')